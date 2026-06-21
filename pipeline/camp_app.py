"""
============================================================================
CAMP Culvert QC - WEB INTERFACE  (Gradio)
============================================================================
A point-and-click interface over the CAMP ML pipeline, for non-technical
users. No terminal needed once it's launched.

  Tab 1  Single Culvert : drag in photos -> see each property prediction.
  Tab 2  Full Dataset QC: run/load the pipeline -> Excel download + an
                          in-browser inspector-vs-model comparison table.

It SHELLS OUT to the existing, tested pipeline scripts (run_inference.py,
dimension_check.py, merge_results.py) so results match the validated
pipeline exactly - nothing is re-implemented here.

RUN (on a machine that has the models, e.g. the GPU node or a laptop with
the models copied):

    export PATH=$HOME/.local/bin:$PATH
    pip install --user gradio openpyxl        # one time
    cd ~/Code/camp_ml/pipeline
    python3 camp_app.py

Gradio prints a local URL (http://127.0.0.1:7860). Open it in a browser.
To let someone on another machine use it, either use SSH port-forwarding
or launch with share=True (see bottom of file) for a temporary public link.
============================================================================
"""

import os
import sys
import glob
import shutil
import tempfile
import subprocess
import datetime

import pandas as pd

PIPELINE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(PIPELINE_DIR)            # ~/Code/camp_ml
CODE_ROOT = os.path.dirname(PROJECT_ROOT)               # ~/Code

def _first_existing(candidates, fallback):
    for c in candidates:
        if c and os.path.exists(c):
            return c
    return fallback

# Predictions/outputs are written by the pipeline into camp_ml/outputs
OUTPUTS_DIR = os.path.join(PROJECT_ROOT, "outputs")

# Inventory CSV and image archive live under ~/Code on this cluster, but be
# robust: allow an env-var override, then check the common locations.
DEFAULT_TABLE = _first_existing([
    os.environ.get("CAMP_TABLE"),
    os.path.join(CODE_ROOT, "nmdot_table_fixed.csv"),     # ~/Code/nmdot_table_fixed.csv  (actual)
    os.path.join(PROJECT_ROOT, "nmdot_table_fixed.csv"),  # ~/Code/camp_ml/...
], os.path.join(CODE_ROOT, "nmdot_table_fixed.csv"))

DEFAULT_IMAGES = _first_existing([
    os.environ.get("CAMP_IMAGES"),
    os.path.join(CODE_ROOT, "all_images"),                # ~/Code/all_images  (actual)
    os.path.join(PROJECT_ROOT, "all_images"),
], os.path.join(CODE_ROOT, "all_images"))
PY = sys.executable

os.makedirs(OUTPUTS_DIR, exist_ok=True)

# Pretty labels for the per-culvert prediction columns
PRED_LABELS = [
    ("Material_Prediction", "Material_Confidence", "Material"),
    ("Shape_Prediction", "Shape_Confidence", "Culvert Shape"),
    ("NumberOfCulverts_Prediction", None, "Number of Culverts"),
    ("Silting_Prediction", "Silting_Confidence", "Silting"),
    ("Corrosion_Prediction", "Corrosion_Confidence", "Corrosion"),
    ("EndSection_Prediction", "EndSection_Confidence", "End Section"),
    ("PhysicalDamage_Prediction", "PhysicalDamage_Confidence", "Physical Damage"),
    ("ChannelCondition_Prediction", "ChannelCondition_Confidence", "Channel Condition"),
    ("ErosionControl_Prediction", "ErosionControl_Confidence", "Erosion Control"),
    ("ChannelType_Prediction", "ChannelType_Confidence", "Channel Type"),
]

# Inventory column -> QC flag column (from merge_results.py) for the compare view
QC_TRIPLES = [
    ("Material", "Material_Prediction", "Material_QC"),
    ("Culvert Shape", "Shape_Prediction", "CulvertShape_QC"),
    ("Number of Culverts", "NumberOfCulverts_Prediction", "NumberofCulverts_QC"),
    ("Silting", "Silting_Prediction", "Silting_QC"),
    ("Corrosion", "Corrosion_Prediction", "Corrosion_QC"),
    ("Outlet End Section Type", "EndSection_Prediction", "EndSection_QC"),
    ("Physical Damage", "PhysicalDamage_Prediction", "PhysicalDamage_QC"),
    ("Channel Condition", "ChannelCondition_Prediction", "ChannelCondition_QC"),
    ("Erosion Control", "ErosionControl_Prediction", "ErosionControl_QC"),
    ("Channel Type", "ChannelType_Prediction", "ChannelType_QC"),
]


# --------------------------------------------------------------------------
# helpers (pure pandas - unit-testable without gradio or the models)
# --------------------------------------------------------------------------
def _newest(pattern):
    matches = sorted(glob.glob(os.path.join(OUTPUTS_DIR, pattern)),
                     key=os.path.getmtime, reverse=True)
    return matches[0] if matches else None


def predictions_to_panel(row):
    """One prediction row (Series) -> tidy DataFrame for display."""
    out = []
    for pred_col, conf_col, label in PRED_LABELS:
        if pred_col not in row or pd.isna(row[pred_col]):
            continue
        pred = str(row[pred_col])
        conf = ""
        if conf_col and conf_col in row and pd.notna(row[conf_col]):
            try:
                conf = f"{float(row[conf_col]) * 100:.0f}%"
            except (ValueError, TypeError):
                conf = str(row[conf_col])
        out.append({"Property": label, "Model prediction": pred, "Confidence": conf})
    return pd.DataFrame(out)


def build_compare_table(review_csv, disagreements_only=False):
    """Load the compact QC review CSV -> a readable comparison DataFrame."""
    df = pd.read_csv(review_csv, dtype=str).fillna("")
    id_col = "Culvert ID" if "Culvert ID" in df.columns else df.columns[0]
    cols = [id_col]
    rename = {id_col: "Culvert ID"}
    for inv, pred, qc in QC_TRIPLES:
        for c in (inv, pred, qc):
            if c in df.columns and c not in cols:
                cols.append(c)
    view = df[cols].copy()
    if disagreements_only:
        qc_cols = [qc for _, _, qc in QC_TRIPLES if qc in view.columns]
        mask = pd.Series(False, index=view.index)
        for qc in qc_cols:
            mask = mask | (view[qc] == "DISAGREE")
        view = view[mask]
    return view


def summarize_qc(review_csv):
    """Counts of AGREE/DISAGREE/etc per property -> summary text."""
    df = pd.read_csv(review_csv, dtype=str).fillna("")
    lines = [f"Total culverts in review file: {len(df)}", ""]
    for _, _, qc in QC_TRIPLES:
        if qc in df.columns:
            vc = df[qc].value_counts()
            agree = int(vc.get("AGREE", 0))
            disagree = int(vc.get("DISAGREE", 0))
            comparable = agree + disagree
            rate = f"{100*agree/comparable:.0f}%" if comparable else "n/a"
            prop = qc.replace("_QC", "")
            lines.append(f"{prop:18s}  agree {agree:5d} | disagree {disagree:5d} "
                         f"| agreement {rate}")
    return "\n".join(lines)


def master_to_excel(master_csv, xlsx_path):
    """Convert master QC CSV to a formatted .xlsx: freeze header, color DISAGREE
    cells red and AGREE green, autofilter."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    df = pd.read_csv(master_csv, dtype=str).fillna("")
    wb = Workbook()
    ws = wb.active
    ws.title = "QC Review"

    header_fill = PatternFill("solid", fgColor="2E5496")
    header_font = Font(color="FFFFFF", bold=True)
    red = PatternFill("solid", fgColor="F8CBAD")
    green = PatternFill("solid", fgColor="C6E0B4")
    amber = PatternFill("solid", fgColor="FFE699")

    # header
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    qc_cols = {c for c in df.columns if c.endswith("_QC") or c.endswith("_Flag")}
    for i, (_, rec) in enumerate(df.iterrows(), start=2):
        for j, col in enumerate(df.columns, start=1):
            val = rec[col]
            cell = ws.cell(row=i, column=j, value=val)
            if col in qc_cols:
                if val == "DISAGREE":
                    cell.fill = red
                elif val == "AGREE":
                    cell.fill = green
                elif val in ("INVALID_DIMENSION", "SPAN_RISE_MISMATCH",
                             "UNSUPPORTED_VALUE"):
                    cell.fill = amber

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df)+1}"
    # reasonable column widths
    for j, col in enumerate(df.columns, start=1):
        width = min(max(len(str(col)) + 2, 12), 32)
        ws.column_dimensions[get_column_letter(j)].width = width

    wb.save(xlsx_path)
    return xlsx_path


# --------------------------------------------------------------------------
# pipeline calls (these need the models present)
# --------------------------------------------------------------------------
def run_single_culvert(image_paths, culvert_id):
    if not image_paths:
        return pd.DataFrame([{"Property": "(no images)", "Model prediction": "", "Confidence": ""}])
    cid = (culvert_id or "culvert").strip().replace("/", "_").replace("\\", "_") or "culvert"
    tmp = tempfile.mkdtemp(prefix="camp_single_")
    img_dir = os.path.join(tmp, "images", cid)
    out_dir = os.path.join(tmp, "out")
    os.makedirs(img_dir)
    os.makedirs(out_dir)
    for p in image_paths:
        if p:
            shutil.copy(p, os.path.join(img_dir, os.path.basename(p)))
    cmd = [PY, os.path.join(PIPELINE_DIR, "run_inference.py"),
           "--images", os.path.join(tmp, "images"), "--output-dir", out_dir]
    res = subprocess.run(cmd, capture_output=True, text=True)
    pred_csv = sorted(glob.glob(os.path.join(out_dir, "camp_predictions_v*.csv")),
                      key=os.path.getmtime, reverse=True)
    if not pred_csv:
        msg = (res.stderr or res.stdout or "inference produced no output").strip()
        return pd.DataFrame([{"Property": "ERROR", "Model prediction": msg[-200:], "Confidence": ""}])
    df = pd.read_csv(pred_csv[0])
    if df.empty:
        return pd.DataFrame([{"Property": "(no culvert detected)", "Model prediction": "", "Confidence": ""}])
    return predictions_to_panel(df.iloc[0])


def run_full_dataset(table_file, use_existing, progress=None):
    table = table_file if table_file else DEFAULT_TABLE
    if not os.path.exists(table):
        return "Inventory CSV not found. Upload it or place nmdot_table_fixed.csv in the project root.", None, None

    if use_existing == "Use latest existing predictions (fast)":
        pred_csv = _newest("camp_predictions_v*.csv")
        if not pred_csv:
            return ("No existing predictions found. Choose 'Run fresh' instead.",
                    None, None)
    else:
        # run inference fresh (slow on CPU)
        r = subprocess.run([PY, os.path.join(PIPELINE_DIR, "run_inference.py"),
                            "--images", DEFAULT_IMAGES, "--output-dir", OUTPUTS_DIR],
                           capture_output=True, text=True)
        pred_csv = _newest("camp_predictions_v*.csv")
        if not pred_csv:
            return ("Inference failed:\n" + (r.stderr or r.stdout)[-400:], None, None)

    # dimensions (optional, best-effort)
    subprocess.run([PY, os.path.join(PIPELINE_DIR, "dimension_check.py"),
                    "--table", table], capture_output=True, text=True)
    dim_csv = _newest("dimension_check_*.csv")

    # merge
    cmd = [PY, os.path.join(PIPELINE_DIR, "merge_results.py"),
           "--table", table, "--predictions", pred_csv, "--output-dir", OUTPUTS_DIR]
    if dim_csv:
        cmd += ["--dimensions", dim_csv]
    r = subprocess.run(cmd, capture_output=True, text=True)

    review_csv = _newest("camp_qc_review_*.csv")
    master_csv = _newest("camp_master_qc_*.csv")
    if not review_csv or not master_csv:
        return ("Merge failed:\n" + (r.stderr or r.stdout)[-400:], None, None)

    stamp = datetime.datetime.now().strftime("%Y_%m_%d")
    xlsx = os.path.join(OUTPUTS_DIR, f"camp_qc_review_{stamp}.xlsx")
    master_to_excel(master_csv, xlsx)

    summary = summarize_qc(review_csv)
    return summary, xlsx, review_csv


# --------------------------------------------------------------------------
# Gradio UI
# --------------------------------------------------------------------------
def build_ui():
    import gradio as gr

    with gr.Blocks(title="CAMP Culvert QC", theme=gr.themes.Soft()) as app:
        gr.Markdown(
            "# CAMP Culvert Quality-Control\n"
            "Automated checks on culvert inspection photos. "
            "Use **Single Culvert** to test a few photos, or **Full Dataset QC** "
            "to review the whole inventory and download an Excel report.")

        with gr.Tab("Single Culvert"):
            gr.Markdown("Drag in one or more photos of the **same** culvert, then click Analyze.")
            with gr.Row():
                with gr.Column(scale=1):
                    imgs = gr.File(file_count="multiple", file_types=["image"],
                                   label="Culvert photos", type="filepath")
                    cid = gr.Textbox(label="Culvert ID (optional, for your reference)")
                    go1 = gr.Button("Analyze culvert", variant="primary")
                with gr.Column(scale=1):
                    out1 = gr.Dataframe(headers=["Property", "Model prediction", "Confidence"],
                                        label="Model predictions", interactive=False, wrap=True)
            go1.click(run_single_culvert, inputs=[imgs, cid], outputs=out1)

        with gr.Tab("Full Dataset QC"):
            gr.Markdown(
                "Run the whole pipeline (or reuse the latest results), then review "
                "model-vs-inspector agreement and download a formatted Excel report.")
            with gr.Row():
                table_in = gr.File(label="Inventory CSV (optional - defaults to nmdot_table_fixed.csv)",
                                   file_types=[".csv"], type="filepath")
                mode = gr.Radio(
                    ["Use latest existing predictions (fast)", "Run fresh inference (slow on CPU)"],
                    value="Use latest existing predictions (fast)", label="Mode")
            go2 = gr.Button("Run QC comparison", variant="primary")
            summary = gr.Textbox(label="Agreement summary", lines=9)
            xlsx_out = gr.File(label="Download Excel report")
            filt = gr.Radio(["All culverts", "Disagreements only"],
                            value="Disagreements only", label="Show")
            table_out = gr.Dataframe(label="Inspector vs. model comparison",
                                     interactive=False, wrap=True)

            state_review = gr.State(None)

            def _run(table_file, m):
                summ, xlsx, review = run_full_dataset(table_file, m)
                if review is None:
                    return summ, None, None, pd.DataFrame()
                tbl = build_compare_table(review, disagreements_only=True)
                return summ, xlsx, review, tbl

            def _refilter(review, choice):
                if not review:
                    return pd.DataFrame()
                return build_compare_table(review, disagreements_only=(choice == "Disagreements only"))

            go2.click(_run, inputs=[table_in, mode],
                      outputs=[summary, xlsx_out, state_review, table_out])
            filt.change(_refilter, inputs=[state_review, filt], outputs=table_out)

        gr.Markdown(
            "<small>Predictions are screening aids. Shape, material, end section and "
            "dimensions are reliable automatic checks; physical damage, erosion control, "
            "silting, corrosion, channel condition and channel type are screening flags "
            "for human review; count is a soft flag (limited by single-end photos). "
            "Channel type's running-water and irrigation classes are provisional pending "
            "more data. Always confirm flagged culverts against the photos.</small>")

    return app


if __name__ == "__main__":
    ui = build_ui()
    # share=False -> local only (use SSH port-forward for remote access).
    # set share=True for a temporary public link (check data-sharing policy first).
    ui.launch(server_name="0.0.0.0", server_port=7860, share=False)
